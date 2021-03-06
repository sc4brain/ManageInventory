# -*- coding: utf-8 -*-
"""
Created on Thu Feb 11 06:59:27 2016

@author: nebula
"""


import sys
import qrcode
import openpyxl as opx
import sqlite3
import urllib

input_filename = '/home/nebula/work/iventory_data/iventory20160209_2.xlsx'
output_dir = '/home/nebula/work/iventory_data/img/'
database_filename = '/home/nebula/work/iventory_data/iventory20160209.sqlite3'


def genQrcode():
    wb = opx.load_workbook(filename=input_filename)
    ws = wb.worksheets[5]
    for i in range(2,100):
        
        code_filename =  ws['B'+str(i)].value.replace(' ', '') + '.png'
        #code_text = u'"Kanzaki", ' + '"' + ws['B'+str(i)].value + '", ' + '"' + ws['C'+str(i)].value + "\""
        code_text = ws['B'+str(i)].value.replace(' ', '')
        code_decoded = urllib.quote(code_text.encode('utf-8'))
        print(code_text + ' -> ' + code_decoded)
        img = qrcode.make(code_decoded)
        img.save(output_dir + code_filename)

def renewTable ():
    conn = sqlite3.connect(database_filename)    
    cur = conn.cursor()

    cur.execute("DROP TABLE iventory")    
    cur.execute("create table iventory(id integer primary key autoincrement, number text, name text, place text, place_detail text, team text, date text, check_person text, stored_date text, comment text, checked bool)")
    conn.commit()
    conn.close()

def genList ():
    wb = opx.load_workbook(filename=input_filename)
    ws = wb.worksheets[5]

    conn = sqlite3.connect(database_filename)    
    cur = conn.cursor()

    for i in range(2, 912):
        record = [ws['B'+str(i)].value, ws['C'+str(i)].value, ws['J'+str(i)].value, 
                 ws['K'+str(i)].value, ws['A'+str(i)].value, ws['Y'+str(i)].value,
                 ws['AA'+str(i)].value]
                 #ws['AA'+str(i)].value + ws['AB'+str(i)].value + ws['AC'+str(i)].value + ws['AG'+str(i)].value]
        record[0] = record[0].replace(" ", "")
        record[1] = record[1].replace("'", "")
        print record
        cur.execute("INSERT INTO iventory(number, name, place, place_detail, team, stored_date, comment, checked) VALUES('%s', '%s', '%s', '%s', '%s', '%s', '%s', %d)" 
        % (record[0], record[1], record[2], record[3], record[4], record[5], record[6], 0))

    
    cur.execute("""SELECT * FROM iventory;""")
    print cur.fetchall()
    conn.commit()
    conn.close()

    #for id, number, name, place in cur.fetchall():
    #    print('%d, %s, %s, %s' % (id, number, name, place))
    '''
    for i in range(2,10):
        
        code_filename =  ws['B'+str(i)].value.replace(' ', '') + '.png'
        code_text = u'"Kanzaki", ' + '"' + ws['B'+str(i)].value + '", ' + '"' + ws['C'+str(i)].value + "\""
        code_decoded = urllib.quote(code_text.encode('utf-8'))
        print(code_text + ' -> ' + code_decoded)
        img = qrcode.make(code_decoded)
        img.save(output_dir + code_filename)
    '''

if __name__ == '__main__':
    #genQrcode()
    renewTable()
    genList()
    